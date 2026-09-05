"""
XLSX_FORMAT - Styling pass applied to output after ANALYZE

Formatting:
    - Frozen header row / leading columns (Code / Song Title / Artist)
    - autofilter & auto-fit column widths
    - Green<yellow<red color scale on difficulties (Official diff, D, RemapDiff, CalcTier)
    - '-1' or missing difficulty placeholder gets a separate white background so it doesn't skew the scale
    - Level (Easy/Medium/Hard/Expert) gets a fixed categorical fill
    - Raw NPS/VPS breakdown and the N/V/COV formula components are hidden, not deleted

progress bar data to Analyze via the `progress` parameter
"""

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule

BODY_FONT = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="A6A6A6")] * 4)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

SCALE_GREEN = "C6EFCE"
SCALE_YELLOW = "FFEB9C"
SCALE_RED = "FFC7CE"

# columns needing '0.00' formatting
FLOAT_COLS = {'aNPS', 'pNPS', 'stdNPS', 'medNPS', 'aVPS', 'pVPS', 'stdVPS', 'medVPS', 'N', 'V', 'COV', 'D'}

# difficulty columns - get a color scale + a bordered box
SCALED_COLS = ['Difficulty', 'D', 'RemapDiff', 'CalcTier']

# raw NPS/VPS + formula pieces, hidden by default but not deleted
DEFAULT_HIDDEN_COLS = ['pNPS', 'aNPS', 'medNPS', 'stdNPS',
                        'pVPS', 'aVPS', 'medVPS', 'stdVPS',
                        'N', 'V', 'COV']

# columns where a blank/NaN value is a real "no data" state, not a gap to fill in - each
# maps to a predicate identifying which raw values in that column count as blank.
# Difficulty's '-1' is analyze.py's placeholder for "no diff_* tag in song.ini"; RemapDiff/
# CalcTier being NaN means "no Expert chart to anchor against for this instrument" (EMHX)
BLANK_PREDICATES = {
    'Difficulty': lambda v: v == -1,
    'RemapDiff': pd.isna,
    'CalcTier': pd.isna,
}

# fixed per-level fill, lighter versions of RB's tier colors - a category, not a gradient
LEVEL_FILL_COLORS = {
    'Easy':   "C6EFCE",
    'Medium': "BDD7EE",
    'Hard':   "FFE5B4",
    'Expert': "FFC7CE",
}
LEVEL_COL = 'Level'

FREEZE_AT = "D2"  # keeps Code / Song Title / Artist visible

# Adds a green-yellow-red color scale for difficulty
def _diff_scale(ws, col_letter, rows):

    if not rows:
        return

    rows = sorted(rows)
    ranges = []
    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        ranges.append((start, prev))
        start = prev = r
    ranges.append((start, prev))

    addr = " ".join(
        f"{col_letter}{a}" if a == b else f"{col_letter}{a}:{col_letter}{b}"
        for a, b in ranges
    )

    ws.conditional_formatting.add(
        addr,
        ColorScaleRule(start_type='min', start_color=SCALE_GREEN,
                        mid_type='percentile', mid_value=50, mid_color=SCALE_YELLOW,
                        end_type='max', end_color=SCALE_RED)
    )


def _named_style(wb, name, **attrs):
    if name not in wb.named_styles:
        style = NamedStyle(name=name)
        for attr, value in attrs.items():
            setattr(style, attr, value)
        wb.add_named_style(style)
    return name


def _body_style(wb, is_scaled, is_float):
    name = f"FW_{'Scaled' if is_scaled else 'Body'}{'Float' if is_float else ''}"
    attrs = {'font': BODY_FONT}
    if is_float:
        attrs['number_format'] = '0.00'
    if is_scaled:
        attrs['border'] = THIN_BORDER
    return _named_style(wb, name, **attrs)


# progress
def style_sheet(ws, df, hidden_cols=DEFAULT_HIDDEN_COLS, scaled_cols=SCALED_COLS,
                 blank_predicates=BLANK_PREDICATES, level_colors=LEVEL_FILL_COLORS,
                 freeze_at=FREEZE_AT, progress=None):
    n_rows, n_cols = df.shape
    columns = list(df.columns)
    scaled_set = set(scaled_cols)
    wb = ws.parent

    # Progress is reported in whole rows
    total_ticks = 2 * n_cols
    ticks_done = 0
    rows_emitted = 0

    def tick():
        nonlocal ticks_done, rows_emitted
        if progress is None:
            return
        ticks_done += 1
        target = int(round(n_rows * ticks_done / total_ticks)) if total_ticks else n_rows
        if target > rows_emitted:
            progress(target - rows_emitted)
            rows_emitted = target

    blank_style = _named_style(wb, 'FW_Blank', font=BODY_FONT, border=THIN_BORDER, fill=WHITE_FILL)
    header_style = _named_style(wb, 'FW_Header', font=HEADER_FONT, alignment=HEADER_ALIGN)
    header_scaled_style = _named_style(wb, 'FW_HeaderScaled', font=HEADER_FONT,
                                        alignment=HEADER_ALIGN, border=THIN_BORDER)

    # one named style per level value, built once and reused across every row
    level_styles = {
        value: _named_style(wb, f'FW_Level_{value}', font=BODY_FONT,
                             fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
        for value, color in level_colors.items()
    }

    ws.freeze_panes = freeze_at
    ws.row_dimensions[1].height = 20

    # style the header + every column's body in one pass, then add the color scale
    for c, col_name in enumerate(columns, start=1):
        is_scaled = col_name in scaled_set
        is_float = col_name in FLOAT_COLS
        is_blank_checked = col_name in blank_predicates
        is_level = col_name == LEVEL_COL

        ws.cell(row=1, column=c).style = header_scaled_style if is_scaled else header_style
        normal_style = _body_style(wb, is_scaled, is_float)

        if is_level:
            values = df[col_name].to_numpy()
            for i, r in enumerate(range(2, n_rows + 2)):
                ws.cell(row=r, column=c).style = level_styles.get(values[i], normal_style)

        elif is_blank_checked:
            predicate = blank_predicates[col_name]
            values = df[col_name].to_numpy()
            real_rows = []
            for i, r in enumerate(range(2, n_rows + 2)):
                cell = ws.cell(row=r, column=c)
                if predicate(values[i]):
                    cell.style = blank_style
                else:
                    cell.style = normal_style
                    real_rows.append(r)
            if is_scaled:
                _diff_scale(ws, get_column_letter(c), real_rows)

        else:
            for r in range(2, n_rows + 2):
                ws.cell(row=r, column=c).style = normal_style
            if is_scaled:
                _diff_scale(ws, get_column_letter(c), list(range(2, n_rows + 2)))

        tick()

    ws.auto_filter.ref = ws.dimensions

    # auto-fit column widths, hiding the raw NPS/VPS/formula columns
    for i, col_name in enumerate(columns, start=1):
        lengths = df[col_name].apply(lambda v: 0 if pd.isna(v) else len(str(v)))
        longest = max(lengths.max(), len(col_name))
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 6), 40)
        if col_name in hidden_cols:
            ws.column_dimensions[get_column_letter(i)].hidden = True
        tick()
